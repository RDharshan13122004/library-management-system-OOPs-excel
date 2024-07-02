import re
from openpyxl import Workbook
from openpyxl import load_workbook
from prettytable import PrettyTable

class user_info:
    def __init__(self,name,age:int,dob,phno,email,addr):

        if not name:
            raise ValueError("missing value")
        
        try:
            assert age >= 12 
        except AssertionError:
            print(f"the age limit is starting with 12")

        try:
            assert re.fullmatch(r"^[0-3][0-9]/[0-1][0-9]/[0-9]{4}$",dob)
        except AssertionError:
            print("Invalid date of birth")

        try:
            assert re.fullmatch(r"^[6-9][0-9]{9}$",phno)
        except AssertionError:
            print("Invalid phone No")

        try:
            assert re.fullmatch(r"^\w+@(\w+\.)?\w+\.(com|org|edu|ac|in|mil)$",email)
        except AssertionError:
            print("Invalid email id")

        wb = load_workbook('ib.xlsx')

        # Select the active worksheet (or specify the sheet name)
        ws = wb['user']

        self.name = name
        self.age = age
        self.dob = dob
        self.email = email
        self.phno = phno
        self.addr = addr
        self.user_id = ws.max_row

        id_ck:bool = False

        wb = load_workbook('ib.xlsx')
        ws = wb['user']

        all_data = []
        for row in ws.iter_rows(values_only=True):
            all_data.append(row[0])

        if self.user_id in all_data:
            id_ck = True
            #print(id_ck)
            
        if id_ck == True:
            print(f"{self.user_id} is already exist retry please")
        else:
            print("successfully data written into sheet: user")
            # Data to append
            new_data = [self.user_id,self.name,self.age,self.dob,self.phno,self.email,self.addr]

            # Append data to the worksheet
            ws.append(new_data)

            # Save the workbook
            wb.save('ib.xlsx')


    def __str__(self):
        return f"name:{self.name}\nage:{self.age}\ndob:{self.dob}\nemail:{self.email}\nphone no:{self.phno}\naddress:{self.addr}\nreg no:{self.user_id}"


class book:
    def __init__(self,title, author, isbn, publisher, year, category, available_copies):
        
        try:
            assert re.fullmatch(r"^[0-9]{4}([0-9a-z]{2})?[0-9a-z]{2}$",isbn)
        except AssertionError:
            print("Invalid Code")

        try:
            assert re.fullmatch(r"^[1-2][0-9]{3}$",year)
        except AssertionError:
            print("Invalid data")

        wb = load_workbook("ib.xlsx")

        ws = wb['books']

        self.book_id = ws.max_row
        self.title = title
        self.author = author
        self.isbn = isbn
        self.publisher = publisher
        self.year = year
        self.category = category
        self.available_copies = available_copies

        bk_id_ck:bool = False
        isbn_ck:bool = False

        wb = load_workbook("ib.xlsx")

        ws = wb['books']
        print(f"Reading data from sheet: books")

        all_detail = []

        for row in ws.iter_rows(min_row=2,values_only=True):
            all_detail.append({"book_id":row[0],"isbn":row[3]})
        """
        for i in all_detail:
            print(f"book id:{i['book_id']},isbn:{i['isbn']}")
        """
        for i in all_detail:
            if str(self.book_id) in str(i["book_id"]):
                bk_id_ck = True
                

            if self.isbn in str(i["isbn"]):
                isbn_ck = True
                
        if bk_id_ck == True:
            print("the book id is already exists")

        if isbn_ck == True:
            print("the isbn code is already exists")

        if bk_id_ck != True and isbn_ck != True:

            new_detail = [self.book_id,self.title,self.author,self.isbn,self.publisher,self.year,self.category,self.available_copies]

            ws.append(new_detail)

            wb.save("ib.xlsx")


    def __str__(self):
        return f"book:{self.book_id}\ntitle:{self.title}\nauthor:{self.author}\nisbn:{self.isbn}\npublisher:{self.publisher}\nyear:{self.year}\ncategory:{self.category}\navailable copies:{self.available_copies}"
    


class transaction:
    def __init__(self,user_id, book_id, borrow_date, due_date, return_date=None, status="borrowed"):
        
        if not user_id:
            raise ValueError("missing value")
        
        if not book_id:
            raise ValueError("missing value")

        try:
            assert re.fullmatch(r"^[0-3][0-9]/[0-1][0-9]/[0-9]{4}$",borrow_date)
        except AssertionError:
            print("Invalid date")

        try:
            assert re.fullmatch(r"^[0-3][0-9]/[0-1][0-9]/[0-9]{4}$",due_date)
        except AssertionError:
            print("Invalid date")

        wb = load_workbook("ib.xlsx")

        wsu = wb["user"]
        wsb = wb["books"]
        wst = wb["transaction"]

        self.transaction_id = wst.max_row
        self.user_id = user_id
        self.book_id = book_id
        self.borrow_date = borrow_date
        self.due_date = due_date
        self.return_date = return_date
        self.status = status

        tran_id_ck:bool = False
        user_id_ck:bool = False
        book_id_ck:bool = False

        user_data = []
        book_data = []
        transaction_data = []

        for row in wsu.iter_rows(min_row=2,values_only=True):
            user_data.append({"user id":row[0]})

        for row in wsb.iter_rows(min_row=2,values_only= True):
            book_data.append({"book id":row[0]})

        for row in wst.iter_rows(min_row=2,values_only=True):
            transaction_data.append({"transaction id":row[0]})
            
        for id  in transaction_data:
            if str(self.transaction_id) in str(id["transaction id"]):
                tran_id_ck = True
                print(tran_id_ck)

        for uid in user_data:
            if self.user_id in str(uid["user id"]):
                user_id_ck = True
                #print(user_id_ck)

        for bid in book_data:
            if self.book_id in str(bid["book id"]):
                book_id_ck = True
                #print(book_id_ck)

        if tran_id_ck == True:
            print("the transaction id is already in exists")
        
        if user_id_ck != True:
            print("there is no user id is the file")

        if book_id_ck != True:
            print("there is no book in this id ")

        if tran_id_ck != True and user_id_ck == True and book_id_ck == True:
            wb = load_workbook("ib.xlsx")

            ws = wb["transaction"]

            new_data = [self.transaction_id,self.user_id,self.book_id,self.borrow_date,self.due_date,self.return_date,self.status]

            ws.append(new_data)

            wb.save("ib.xlsx")
                
            print(f"transaction ID:{self.transaction_id}\nuser ID:{self.user_id}\nbook ID:{self.book_id}\nborrowdata:{self.borrow_date}\ndue data:{self.due_date}\nreturn date:{self.return_date}\nstatus:{self.status}")
        

class Library:

    def __init__(self,filename = "ib.xlsx"):

        self.filename = filename
        try:
            self.wb = load_workbook(self.filename)
        except FileNotFoundError:
            self.wb = Workbook()
            self.setup_sheets()

        self.user_sheets = self.wb["user"]
        self.book_sheets = self.wb["books"]
        self.transaction_sheet = self.wb["transaction"]
        self.save_wb()

    def setup_sheets(self):
        self.user_sheets = self.wb.create_sheet("user")
        self.user_sheets.append(['user id','name','age','dob','phone no','email','address'])
        self.book_sheets = self.wb.create_sheet("books")
        self.book_sheets.append(['book id','title','author','isbn','publisher','year','category','available copies'])
        self.transaction = self.wb.create_sheet("transaction")
        self.transaction.append(['transaction id','user id','book id','borrow date','due date','return date','status'])

    def save_wb(self):
        self.wb.save(self.filename)

    #user management

    def user_reg(self):
        name = input("enter the name:").strip().lower()
        age = int(input("enter the age:"))
        dob = input("enter the dob:")
        phno = input("enter the phone:")
        email = input("enter the email:")
        addr = input("ente the address:")

        user1 = user_info(name,age,dob,phno,email,addr)   
        print(user1)

    def view_user(self):

        print(f"Data from sheet:user")
        table = PrettyTable()

        # Extract headers from the first row
        headers = [cell.value for cell in self.user_sheets[1]]
        table.field_names = headers

        # Extract all rows from the worksheet
        for row in self.user_sheets.iter_rows(min_row=2, values_only=True):
            table.add_row(row)

        # Print the table
        print(table)
    
    def get_user(self):
        names = input("enter your name:").strip().lower()
        phoneno = input("enter your phone number:")
        if not names:
            raise ValueError("missing value")
        
        if not re.fullmatch(r"^[6-9][0-9]{9}$",phoneno):
            raise ValueError("Invalid phone No")

        
        user_data = []
        for u in self.user_sheets.iter_rows(min_row=2,values_only=True):
            user_data.append({"user id":u[0],"user":u[1],"age":u[2],"dob":u[3],"phno":u[4],"email":u[5],"addr":u[6]})

        ck_nam:bool = False
        ck_phno:bool = False
        
        for r in user_data:
            if names in str(r['user']):
                ck_nam = True
                if phoneno in str(r['phno']):
                    ck_phno = True
                    print(f"Name:{r['user']}\nUser id:{r['user id']}\nage:{r['age']}\ndob:{r['dob']}\n{r['phno']}\nemail:{r['email']}\naddress:{r['addr']}")
                    
        if ck_nam != True:
            print("Invalid user name")

        if ck_phno != True:
            print("Invalid phone no")

    #books management

    def books(self):

        book_title = input("enter the book title:").strip().lower()
        book_author = input("enter the author:").strip().lower()
        book_isbn = input("enter the isbn code:")
        book_publisher = input("enter the publisher:").strip().lower()
        in_year = input("enter the year:")
        book_category = input("enter the category of the book:").strip().lower()
        ava_copies = input("enter the avaiable no.of copies:")

        user2 = book(book_title,book_author,book_isbn,book_publisher,in_year,book_category,ava_copies)
        print(user2)

    def view_books(self):

        table = PrettyTable()

        header = [cell.value for cell in self.book_sheets[1]] 
        table.field_names = header

        for row in self.book_sheets.iter_rows(min_row=2,values_only=True):
            table.add_row(row)

        print(table)

    def get_book_title(self):
        book_title = input("enter the book title:").strip().lower()

        all_detail = []

        for row in self.book_sheets.iter_rows(min_row=2,values_only=True):
            all_detail.append({"book id":row[0],"title":row[1],"author":row[2],"isbn":row[3],"publisher":row[4],"year":row[5],"category":row[6],"available copies":row[7]}) 

        for books in all_detail:
            if book_title in str(books["title"]):
                print(f"book id:{books['book id']}\ntitle:{books['title']}\nauthor:{books['author']}\nisbn:{books['isbn']}\npublisher:{books['publisher']}\nyear:{books['year']}\ncategory:{books['category']}\navailable copies:{books['available copies']}\n")   

    def get_book_author(self):
        book_author = input("enter the author of the book:").strip().lower()
    
        all_detail = []

        for row in self.book_sheets.iter_rows(min_row=2,values_only=True):
            all_detail.append({"book id":row[0],"title":row[1],"author":row[2],"isbn":row[3],"publisher":row[4],"year":row[5],"category":row[6],"available copies":row[7]}) 

        for books in all_detail:
            if book_author in str(books["author"]):
                print(f"book id:{books['book id']}\ntitle:{books['title']}\nauthor:{books['author']}\nisbn:{books['isbn']}\npublisher:{books['publisher']}\nyear:{books['year']}\ncategory:{books['category']}\navailable copies:{books['available copies']}\n")


    def get_book_isbn(self):
        book_isbn = input("enter the isbn code of the book:").strip().lower()
        
        all_detail = []

        for row in self.book_sheets.iter_rows(min_row=2,values_only=True):
            all_detail.append({"book id":row[0],"title":row[1],"author":row[2],"isbn":row[3],"publisher":row[4],"year":row[5],"category":row[6],"available copies":row[7]}) 

        for books in all_detail:
            if book_isbn in str(books["isbn"]):
                print(f"book id:{books['book id']}\ntitle:{books['title']}\nauthor:{books['author']}\nisbn:{books['isbn']}\npublisher:{books['publisher']}\nyear:{books['year']}\ncategory:{books['category']}\navailable copies':{books['available copies']}\n")

    def get_book_publisher(self):
        book_publisher = input("enter the publisher of the book:").strip().lower()

        all_detail = []

        for row in self.book_sheets.iter_rows(min_row=2,values_only=True):
            all_detail.append({"book id":row[0],"title":row[1],"author":row[2],"isbn":row[3],"publisher":row[4],"year":row[5],"category":row[6],"available copies":row[7]}) 

        for books in all_detail:
            if book_publisher in str(books["publisher"]):
                print(f"book id:{books['book id']}\ntitle:{books['title']}\nauthor:{books['author']}\nisbn:{books['isbn']}\npublisher:{books['publisher']}\nyear:{books['year']}\ncategory:{books['category']}\navailable copies':{books['available copies']}\n")

    def get_book_category(self):
    
        all_detail = []

        for row in self.book_sheets.iter_rows(min_row=2,values_only=True):
            all_detail.append({"book id":row[0],"title":row[1],"author":row[2],"isbn":row[3],"publisher":row[4],"year":row[5],"category":row[6],"available copies":row[7]}) 

        cat = set()
        for a_d in all_detail:
            cat.add(a_d["category"])

        table = PrettyTable()

        table.field_names = ["Category"]
        for row in sorted(cat):
            table.add_row([row])

        print(table)

        book_category = input("enter the category of the book:").strip().lower()

        for books in all_detail:
            if book_category in str(books["category"]):
                print(f"book id:{books['book id']}\ntitle:{books['title']}\nauthor:{books['author']}\nisbn:{books['isbn']}\npublisher:{books['publisher']}\nyear:{books['year']}\ncategory:{books['category']}\navailable copies:{books['available copies']}\n")


    #transaction management

    def transactions(self):

        u_id = input("enter the user id:")
        b_id = input("enter the book id:")
        borrow_date = input("enter the borrow data:")
        due_date = input("enter the due date:")

        all_book_list = []
        user_id_list = []

        book_ava:bool = False
        user_id_ck:bool = False

        for book in self.user_sheets.iter_rows(min_row=2,values_only=True):
            user_id_list.append({"user id":book[0]})

        for book in self.book_sheets.iter_rows(min_row=2,values_only=True):
            all_book_list.append({"book id":book[0],"available copies":book[7]})

        for u in user_id_list:
            if u_id in str(u["user id"]):
                user_id_ck = True
                for a in all_book_list:
                    if b_id in str(a["book id"]):
                        if int(a["available copies"]) > 0:
                            #print(user)
                            self.book_ava_cop_der(b_id)
                            transaction(u_id,b_id,borrow_date,due_date) 
                            book_ava = True

        if not user_id_ck:
            print("the user id is not exists")

        if book_ava:
            print("the book is available in the library and successfully updated in transaction sheet")
            #self.view_transaction()

        if not book_ava:
            print("the book is not available in the library")
        

    def book_ava_cop_der(self,b_id):
        bk_id = b_id
        result:bool = False
        for row in self.book_sheets.iter_rows(min_row=2):
            if bk_id == str(row[0].value):
                #print(row[0].value)
                available_copies = row[7].value
                if available_copies is not None and available_copies > 0:
                    #print(available_copies)
                    row[7].value = int(available_copies) - 1
                    result = True
                    #self.save_wb()
                    #print(result)
                    break

        if result == True:
            self.save_wb()
        else:
            print("failed to update book availability.")

    def view_transaction(self):

        table = PrettyTable()

        header = [cell.value for cell in self.transaction_sheet[1]]
        table.field_names = header

        for row in self.transaction_sheet.iter_rows(min_row=2,values_only=True):
            table.add_row(row)

        print(table)

    def return_book(self):
        
        t_id = input("Enter the transaction id:").strip()
        return_date = input("Enter the return date:")

        if not re.fullmatch(r"^[0-3][0-9]/[0-1][0-9]/[0-9]{4}$",return_date):
            raise ValueError("Invalid date format")

        transaction_found = False

        for row in self.transaction_sheet.iter_rows(min_row=2):
            if row[0].value == int(t_id) and row[6].value == "borrowed":
                row[5].value = return_date
                row[6].value = "returned"
                transaction_found = True

                # Update book availability
                book_id = row[2].value
                for book_row in self.book_sheets.iter_rows(min_row=2):
                    if str(book_row[0].value) == book_id:
                        book_row[7].value += 1
                        break
                self.save_wb()
                print("Book returned successfully and availability updated.")
                break

        if not transaction_found:
            print("Transaction not found or book already returned.")
